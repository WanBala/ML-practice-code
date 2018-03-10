# -*- coding: utf-8 -*-
"""
Created on Thu Jan 18 18:39:27 2018

@author: YU-TING

此程式是由Excel讀出資料
並以列為單位排除缺失值

如此可以從Excel中讀出完整無破損的資料進行機器學習






"""





#交叉驗證工具

from openpyxl import load_workbook
import numpy as np

def Cross_check(file_name,index_row,index_column):      #(檔案名稱, 資料從哪行開始,資料從哪欄開始)
    Data=load_workbook(file_name)
    Data_Sheet=Data.active
    Total_Rows=Data_Sheet.max_row
    Total_Columns=Data_Sheet.max_column
    invalid_row=0
    for row in range(index_row,Total_Rows+1):
        for column in range(index_column,Total_Columns+1):
            Eng=chr(column+64)
            if(type(Data_Sheet[Eng+str(row)].value)==float or type(Data_Sheet[Eng+str(row)].value)==int):
                pass
            else:
                invalid_row+=1
                break
    valid_row=Total_Rows-invalid_row
    throw_rate=invalid_row/(Total_Rows-index_row+1)
    return valid_row,throw_rate       #回傳有效樣本量,無效樣本率

def Valid_get(file_name,index_row,index_column): #(檔案名稱,資料從哪行開始,資料從哪欄開始)
    Data=load_workbook(file_name)
    Data_Sheet=Data.active
    Total_Rows=Data_Sheet.max_row
    Total_Columns=Data_Sheet.max_column
    array=[]
    for row in range(index_row,Total_Rows+1):
        cache=[]
        for column in range(index_column,Total_Columns+1):
            Fake=0
            Eng=chr(column+64)
            if(type(Data_Sheet[Eng+str(row)].value)==float or type(Data_Sheet[Eng+str(row)].value)==int):
                cache.append(Data_Sheet[Eng+str(row)].value)    
            else:
                Fake=1
                break
        if(Fake==1):
            continue
        else:
            array.append(cache)
    
    NumArray=np.array(array)
    return NumArray                    #回傳Numpy矩陣
#----------------------------------------------------

